# This is a Python script used to generate a series of JSON files from the cmn_configuration_data spreadsheet to be imported into NetBox
import argparse
from openpyxl import load_workbook
import json
import re

def create_fresh_device_import_dict():

    device_import_format = {
        "name":"",
        "manufacturer":"",
        "device_type":"",
        "status":"",
        "site":"",
        "role":"",
        "cf_management_vrf":"",
        "cf_bgp_asn":"",
        "cf_ptp_domain":"",
        "cf_ptp_p1":"",
        "cf_ptp_p2":"",
        "cf_dhcp_server_en":""
    }

    return device_import_format

def create_fresh_int_import_dict():

        int_import_format = {
            "device":"",
            "name":"",
            "type":"",
            "label":"",
            "speed":"",
            "duplex":"",
            "enabled":True,
            "mode":"",
            "vrf":"",
            "parent":"",
            "mgmt_only":"",
            "description":"",
            "cf_pim_enabled":False,
            "cf_multicast_boundary":"",
            "cf_acl_in":"",
            "cf_ptp_profile":"",
            "cf_ptp_master":""
        }

        return int_import_format

def create_fresh_ip_import_dict():

    ip_import_format = {
        "address":"",
        "status":"",
        "role":"",
        "device":"",
        "description":""
    }

    return ip_import_format

def get_site_name(hostname):

    # This function returns the site name from the hostname of the device    

    if re.match("EC-MA", hostname) is not None:
        return "Energy Centre"
    elif re.match("WN-MW", hostname) is not None:
        return "Wood Norton"
    elif re.match("MR-CAA", hostname) is not None:
        return "Media City"
    elif re.match("GW-IDR", hostname) is not None:
        return "Pacific Quays"
    else:
        return ""

def set_int_speed(int_speed):
    
    # This function converts the interface speed to kbps

    if (int_speed == "100gfull"):
        return "100000000"
    elif (int_speed == "10000full"):
        return "10000000"
    elif (int_speed == "1000full"):
        return "1000000"

def set_int_type(hostname, int_name, device_ws):

    # Only parent interfaces are passed to this function, subinterfaces have the 'virtual' type used instead

    switch_type = ""
    
    # Get switch type from hostname
    for row in device_ws.iter_rows(min_row=2, max_row=device_ws.max_row, values_only=True):
        if row[0] == hostname:
            switch_type = row[2]
    
    # Get interface port number
    int_num = int_name.replace("Ethernet", "")
    if re.search(r'/\d+$', int_num): # remove /x from interface name
        int_num = int_num[:-2]
    int_num = int(int_num)

    if (switch_type == "7280CR3-96") or (switch_type == "7280CR2A-30") or (switch_type == "CCS-710P-12"): # All interfaces are 100G - QSFP
        return "100gbase-x-qsfp28"

    elif switch_type == "7280SR2-48YC6":
        if (int_num > 0 and int_num <= 48):
            return "25gbase-x-sfp28"
        elif (int_num > 48 and int_num <=54):
            return "100gbase-x-qsfp28"

    elif switch_type == "7010T-48":
        if (int_num > 0 and int_num <= 48):
            return "1000base-t"
        elif (int_num > 48 and int_num <=52):
            return "10gbase-x-sfpp"

    elif (switch_type == "7020TR-48") or (switch_type == "DCS-7020TR-48"):
        if (int_num > 0 and int_num <= 48):
            return "1000base-t"
        elif (int_num > 48 and int_num <=54):
            return "10gbase-x-sfpp"

    elif (switch_type == "7020SR-24C2") or (switch_type == "DCS-7020SR-24C2"):
        if (int_num > 0 and int_num <= 24):
            return "10gbase-x-sfpp"
        elif (int_num > 24 and int_num <=26):
            return "100gbase-x-qsfp28"
    
    else:
        print(f"Error cannot find port interface type for {hostname} {int_name} of switch type - {switch_type}.")


def generate_device_json(device_ws, grey_net):

    # This function generates all the devices to be imported into NetBox

    device_json_list = []
    offset = 0
    if grey_net == True: offset = 1 # Grey network has an extra loopback column in the device sheet which offsets all subsequent columns

    for row in device_ws.iter_rows(min_row=2, max_row=device_ws.max_row, values_only=True): # min_row 0 & 1 both include headers

        device_import_format = create_fresh_device_import_dict()

        device_import_format["name"] = row[0]
        device_import_format["manufacturer"] = "Arista"
        device_import_format["device_type"] = row[2]
        device_import_format["status"] = "active"
        device_import_format["site"] = get_site_name(row[0])

        # Rename switch role to match existing roles in NetBox
        ws_role = row[3]
        if ws_role == "leaf":
            device_import_format["role"] = "Leaf"
        elif ws_role == "spine":
            device_import_format["role"] = "Spine"
        elif ws_role == "timing":
            device_import_format["role"] = "Timing"
        elif ws_role == "tapagg":
            device_import_format["role"] = "TapAgg"

        device_import_format["cf_management_vrf"] = row[7]
        device_import_format["cf_bgp_asn"] = row[9 + offset]
        device_import_format["cf_ptp_domain"] = row[13 + offset]
        device_import_format["cf_ptp_p1"] = row[14 + offset]
        device_import_format["cf_ptp_p2"] = row[15 + offset]

        
        if (row[16 + offset] == "enabled"):
            device_import_format["cf_dhcp_server_en"] = True
        else:
            device_import_format["cf_dhcp_server_en"] = False

        device_json_list.append(dict(device_import_format))

    return(device_json_list)

# Generate interface JSON

def generate_int_json(interface_ws, device_ws, grey_net):
    
    # This function generates all the port, loopback and management interfaces to be imported into NetBox

    int_json_list = []

    for row in interface_ws.iter_rows(min_row=2, max_row=interface_ws.max_row, values_only=True): # min_row 0 & 1 both include headers

        int_import_format = create_fresh_int_import_dict()

        if (row[2] == "enabled"):    
            int_import_format["device"] = row[0] # device hostname
            int_import_format["name"] = f"{row[0]}-{row[1].replace('thernet', '')}" # shorten interface name to hostname + -Ex/y
            # int_import_format["name"] = row[1]

            # Interface type will vary depending on the switch port type and whether the interface is a sub interface or not
            # The presence of a point '.' in the interface name indicates the interface is a sub interface
            # Sub interfaces will be of type 'virtual' and will have a parent interface
            if re.search(r'\.', row[1]): # subinterface found
                int_import_format["type"] = "virtual" # interface is of type 'virtual'
                par_int = f"{row[0]}-{(row[1].replace('thernet', '')).split('.')[0]}" # set the parent interface of the subinterface
                int_import_format["parent"] = par_int
                int_import_format["mode"] = "access"
            
            else: # 'normal' interface
                int_import_format["type"] = set_int_type(row[0], row[1], device_ws) # pass hostname, interface speed and device worksheet to return the interface type
                int_import_format["parent"] = ""

            int_import_format["label"] = row[1]
            int_import_format["speed"] = set_int_speed(row[3])
            int_import_format["description"] = f"Connection to {row[6]} {row[7]}" # description is the device and device interface this interface connects to
            int_import_format["duplex"] = "full"

            # Red & Blue Network spreadsheets have an extra 3 columns for DHCP addressing
            if (grey_net == True): int_import_format["vrf"] = row[16] # Grey
            else: int_import_format["vrf"] = row[19] # Red & Blue
            
            if row[10] == "enabled": int_import_format["cf_pim_enabled"] = True
            int_import_format["cf_multicast_boundary"] = row[11]
            int_import_format["cf_acl_in"] = row[12]
            int_import_format["cf_ptp_profile"] = row[13]
            
            if row[14] == "enabled": int_import_format["ptp_master"] = True

            int_json_list.append(dict(int_import_format))
        
        else:
            pass # only import enabled interfaces
    
    # Generate management interfaces

    for row in device_ws.iter_rows(min_row=2, max_row=device_ws.max_row, values_only=True): # min_row 0 & 1 both include headers

        int_import_format = create_fresh_int_import_dict()

        int_import_format["device"] = row[0]
        int_import_format["name"] = f"{row[0]} Management"
        int_import_format["type"] = "1000base-t"
        int_import_format["label"] = f"Management"
        int_import_format["mgmt_only"] = True
        int_import_format["description"] = f"{row[0]} Management Interface"

        int_json_list.append(dict(int_import_format))
    
    # Generate loopback interfaces

    for row in device_ws.iter_rows(min_row=2, max_row=device_ws.max_row, values_only=True): # min_row 0 & 1 both include headers

        int_import_format = create_fresh_int_import_dict()

        int_import_format["device"] = row[0]
        int_import_format["name"] = f"{row[0]} Loopback"
        int_import_format["type"] = "virtual"
        int_import_format["label"] = f"Loopback"
        int_import_format["description"] = f"{row[0]} Loopback Interface"

        int_json_list.append(dict(int_import_format))

    return(int_json_list)

# Generate IP address JSON

def generate_ip_json(interface_ws, device_ws, grey_net):
    
    # This function generates all the device port, management and loopback interface IP addresses to be imported into NetBox

    ip_json_list = []

    # Generate port addresses

    for row in interface_ws.iter_rows(min_row=2, max_row=interface_ws.max_row, values_only=True): # min_row 0 & 1 both include headers
        
        ip_import_format = create_fresh_ip_import_dict()

        if ((row[8] != None) and (row[9] != None)):
            
            ip_import_format["address"] = f"{row[8]}/{row[9]}"
            if (row[2] == "enabled"): ip_import_format["status"] = "active"
            else: ip_import_format["status"] = "reserved"
            ip_import_format["device"] = row[0]
            ip_import_format["interface"] = f"{row[0]}-{row[1].replace('thernet', '')}"
            ip_import_format["description"] = f"{row[0]} - {row[1]} interface address"

            ip_json_list.append(dict(ip_import_format))

    # Generate management addresses

    for row in device_ws.iter_rows(min_row=2, max_row=device_ws.max_row, values_only=True): # min_row 0 & 1 both include headers
        
        ip_import_format = create_fresh_ip_import_dict()
       
        ip_import_format["address"] = f"{row[4]}/{row[5]}"
        ip_import_format["status"] = "active"
        ip_import_format["role"] = ""
        ip_import_format["device"] = row[0]
        ip_import_format["description"] = f"{row[0]} management address"

        ip_json_list.append(dict(ip_import_format))

    # Generate loopback addresses

    for row in device_ws.iter_rows(min_row=2, max_row=device_ws.max_row, values_only=True): # min_row 0 & 1 both include headers
        
        ip_import_format = create_fresh_ip_import_dict()

        if (row[8] == None): 
            ip_import_format["address"] = f"{row[9]}/24" # Grey switches in MR + PQ use Loopback2 addresses?
        else:
            ip_import_format["address"] = f"{row[8]}/24"
        ip_import_format["status"] = "active"
        ip_import_format["role"] = "loopback"
        ip_import_format["device"] = row[0]
        ip_import_format["description"] = f"{row[0]} loopback address"

        ip_json_list.append(dict(ip_import_format))
    
    return(ip_json_list)

def main():

    parser = argparse.ArgumentParser(
        description = "Generate Netbox import files from CMN Configuration spreadsheets."
        )

    # Define arguments required to run script
    parser.add_argument(
        "-f",
        help = "Specify a configuration spreadsheet",
        dest = "file",
        required = True
    )

    # Get passed arguments
    args = parser.parse_args()

    # Get cmn file
    cmn_file = load_workbook(args.file)
    
    # Check Red & Blue or Grey Network, each spreadsheet holds a different number of columns
    site = ""
    grey_net = False

    site_pattern = r'sco'
    network_pattern = r'grey'
    if re.search(site_pattern, args.file):
        site = "sco"
    else:
        site = "nco"
    if re.search(network_pattern, args.file):
        grey_net = True

    device_ws = cmn_file["device"]
    interface_ws = cmn_file["interface"]

    # Create a JSON file from the device worksheet to import into the devices page on NetBox  
    device_json_list = generate_device_json(device_ws, grey_net)

    # Generate device interface list
    int_json_list = generate_int_json(interface_ws, device_ws, grey_net) # device worksheet required to get interface type from switch model

    # Generate interface IP address list
    ip_json_list = generate_ip_json(interface_ws, device_ws, grey_net)

    # Set prefix of save file names:
    network = "red_blue"
    if grey_net == True: network = "grey"

    # Write data to JSON files
    path = f"C:\\Users\\caddem01\\OneDrive - BBC\\BBC\\Code\\Python Code\\netbox-cmn-import-script\\uploads\\{network}\\"

    # Save the device import JSON
    save_file_name = f"{site}_{network}_device_import.json"
    with open(path + save_file_name, "w") as open_file:
        json.dump(device_json_list, open_file, indent=2) # pretty JSON file

    # Save the device interface import JSON
    save_file_name = f"{site}_{network}_int_import.json"
    with open(path + save_file_name, "w") as open_file:
        json.dump(int_json_list, open_file, indent=2) # pretty JSON file

    # Save the interface IP address import JSON
    save_file_name = f"{site}_{network}_ip_import.json"
    with open(path + save_file_name, "w") as open_file:
        json.dump(ip_json_list, open_file, indent=2) # pretty JSON file

if __name__ == "__main__":
    main()